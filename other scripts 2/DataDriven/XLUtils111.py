import openpyxl

class XLUtils:
    def __init__(self, file):
        self.file = file
        self.workbook = openpyxl.load_workbook(self.file)

    def getRowCount(self, sheetName):
        sheet = self.workbook[sheetName]
        return sheet.max_row

    def getColumnCount(self, sheetName):
        sheet = self.workbook[sheetName]
        return sheet.max_column

    def readData(self, sheetName, rownum, columnno):
        sheet = self.workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(self, sheetName, rownum, columnno, data):
        sheet = self.workbook[sheetName]
        sheet.cell(row=rownum, column=columnno, value=data)
        self.workbook.save(self.file)