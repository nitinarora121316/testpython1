import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Open the excel file
# workbook = openpyxl.load_workbook("locators.xlsx")
workbook = openpyxl.load_workbook("C:\\Users\\asus\\Desktop\\locators.xlsx")

sheet = workbook.active

# Fetch the locators from the excel file
search_bar_locator = sheet.cell(row=2, column=2).value

# Initialize a new browser instance
driver = webdriver.Chrome(ChromeDriverManager().install())

# Navigate to a website
driver.get("https://ap.bayatree.com")

wait = WebDriverWait(driver, 10)

# Find the search bar element using the locator from the excel file
search_bar = driver.find_element(By.XPATH, search_bar_locator)
search_bar.send_keys("selenium python")
time.sleep(5)

# Wait for the page to load
driver.implicitly_wait(10)

# Print the title of the page
print(driver.title)

# Close the browser
driver.quit()
